import requests
import time
import json
from openpyxl import Workbook
from settings import *

price_list = dict()


def write_html(data, name, format_p='.html'):
    p1 = ps(f'{name}{format_p}')
    with open(p1, 'w', encoding='utf-8') as file:
        file.write(data.text)


def parsing_id_product_from_html(name):
    """
    Открывает файл в текущей папке и конвертирует из json в словарь
    """
    with open(name, 'r', encoding='utf-8') as file:
        data = file.read()
        data = json.loads(data)
        total = data['body']['total']
        products = ','.join(data['body']['products'])
    return [products, total]


def get_urls(urls, params=None, data=None, cookies=None, req_type='get', save=None):
    time.sleep(1)
    response1 = ''
    for i in range(10):
        try:
            if req_type == 'get':
                response1 = requests.get(urls, params=params, data=data, headers=HEADERS, cookies=cookies, timeout=10)
            elif req_type == 'post':
                response1 = requests.post(urls, params=params, data=data, headers=HEADERS, cookies=cookies, timeout=10)
            print('status_code', response1.status_code, req_type, urls)
            if 200 <= response1.status_code < 300:
                break
            else:
                time.sleep(4)
        except Exception as e:
            print('Ошибка: \n', e)
            time.sleep(3)

    if save is not None:
        write_html(response1, save)
    return response1


def main():
    i = 0
    res = get_urls(urls=url_start)
    cookies = dict(res.cookies.items())
    for pg in product_groups:
        count = 0
        while True:
            params = {'categoryId': pg, 'offset': count, 'limit': '72', 'doTranslit': 'true'}
            get_urls(urls=url_id, params=params, cookies=cookies, save=f'url_id/url_id{i}')

            normalization_path = ps(f'url_id/url_id{i}.html')
            product_code = parsing_id_product_from_html(normalization_path)

            products = product_code[0].split(',')
            data = {"productIds": products, "media": 'true', "categories": 'true', "availability": 'true'}
            get_urls(urls=url_product, data=data, cookies=cookies, req_type='post', save=f'url_product/url_product{i}')

            params = {'productIds': product_code[0]}
            get_urls(urls=url_prices, params=params, cookies=cookies, save=f'url_prices/url_prices{i}')
            get_urls(urls=url_statuses, params=params, cookies=cookies, save=f'url_statuses/url_statuses{i}')

            i += 1
            count += 72
            print(f'Сканировано --{count}--из--{product_code[1]}--, итерация №{i}')
            counter = product_code[1]
            if count > counter:
                break


def open_convert_html2(name, type_data):
    for file_folder in os.listdir(name):
        if file_folder.endswith(".html"):
            file_folder = os.path.normpath(f'{name}/{file_folder}')
            with open(file_folder, 'r', encoding='utf-8') as file:
                data = file.read()
                data = json.loads(data)
                if type_data == 'product':
                    for i1 in data['body']['products']:
                        price_list[i1['productId']] = [i1['name'], '', '', '']
                elif type_data == 'prices':
                    for i2 in data['body']['materialPrices']:
                        base_price = i2['price']['basePrice']
                        sale_price = i2['price']['salePrice']
                        price_list[i2['productId']][1] = base_price
                        price_list[i2['productId']][2] = sale_price
                elif type_data == 'statuses':
                    for i2 in data['body']['statuses']:
                        product_status = i2['productStatus']
                        price_list[i2['productId']][3] = product_status


def ps(name):
    head, tail = os.path.split(__file__)
    p1 = os.path.normpath(f'{head}/data/{name}')
    return p1


def file_cleaner():
    file_folder_list = {'url_id': ['url_id', 'url_id/Архив'],
                        'url_prices': ['url_prices', 'url_prices/Архив'],
                        'url_product': ['url_product', 'url_product/Архив'],
                        'url_statuses': ['url_statuses', 'url_statuses/Архив']}

    for i in file_folder_list.values():
        u1 = ps(i[0])
        u2 = ps(i[1])
        for file_folder in os.listdir(u1):
            if file_folder.endswith(".html"):
                head, tail = os.path.split(file_folder)
                file_folder1 = os.path.normpath(f'{u1}/{file_folder}')
                file_folder2 = os.path.normpath(f'{u2}/{tail}')
                os.replace(src=file_folder1, dst=file_folder2)


def writing_file_excel():
    """
    Сохраняем данные в формате excel
    """
    wb = Workbook()
    ws = wb.active

    i = 0
    for pr in price_list.values():
        if pr[3] == 'Available':
            ws.cell(row=i + 1, column=1, value=pr[0])
            ws.cell(row=i + 1, column=2, value=pr[1])
            ws.cell(row=i + 1, column=3, value=pr[2])
            ws.cell(row=i + 1, column=4, value=pr[3])
            i += 1

    wb.save(filename=path_file_save)
